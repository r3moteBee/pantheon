"""Server-side preview generation for binary artifact types.

All Office and PDF previews are read-only. The Settings sandbox card
exposes which renderers are available based on what's installed.

Cache layout: data/previews/<artifact_id>/<version_id>/<asset>
"""
from __future__ import annotations

import asyncio
import logging
import re
import shutil
from pathlib import Path
from typing import Any

from config import get_settings

logger = logging.getLogger(__name__)


_SVG_SCRIPT_RE = re.compile(r"<script\b[^>]*>.*?</script>", re.IGNORECASE | re.DOTALL)
_SVG_ON_ATTR_RE = re.compile(r'\son[a-z]+\s*=\s*"[^"]*"', re.IGNORECASE)
_SVG_ON_ATTR_RE_SQ = re.compile(r"\son[a-z]+\s*=\s*'[^']*'", re.IGNORECASE)
_SVG_HREF_RE = re.compile(r'(xlink:href|href)\s*=\s*"(?!data:)[^"]*"', re.IGNORECASE)


def sanitize_svg(svg: str) -> str:
    """Strip <script>, on* handlers, and external xlink:href URIs."""
    svg = _SVG_SCRIPT_RE.sub("", svg)
    svg = _SVG_ON_ATTR_RE.sub("", svg)
    svg = _SVG_ON_ATTR_RE_SQ.sub("", svg)
    svg = _SVG_HREF_RE.sub("", svg)
    return svg


def preview_dir(artifact_id: str, version_id: str) -> Path:
    s = get_settings()
    p = s.data_dir / "previews" / artifact_id / version_id
    p.mkdir(parents=True, exist_ok=True)
    return p


def invalidate_for_artifact(artifact_id: str) -> None:
    s = get_settings()
    p = s.data_dir / "previews" / artifact_id
    if p.exists():
        shutil.rmtree(p, ignore_errors=True)


async def render_preview(artifact: dict[str, Any], blob: bytes | None) -> dict[str, Any]:
    """Return a dict the API sends back to the frontend.

    Shape:
      {type: "html"|"image"|"sheet"|"pdf"|"svg"|"text"|"unsupported",
       content?: str, url?: str, sheets?: list, ...}
    """
    ct = artifact["content_type"].lower()

    # Text / chat-export → frontend renders the content directly
    from artifacts.store import is_text_type
    if is_text_type(ct):
        if ct == "image/svg+xml":
            return {"type": "svg", "content": sanitize_svg(artifact.get("content") or "")}
        return {"type": "text", "content": artifact.get("content") or ""}

    if blob is None:
        return {"type": "unsupported", "reason": "no blob loaded"}

    if ct.startswith("image/"):
        # Frontend uses the raw endpoint via <img>
        return {"type": "image", "url": f"/api/artifacts/{artifact['id']}/raw"}

    if ct == "application/pdf":
        return {"type": "pdf", "url": f"/api/artifacts/{artifact['id']}/raw"}

    if ct.endswith("wordprocessingml.document"):
        return await _render_docx(artifact, blob)
    if ct.endswith("spreadsheetml.sheet"):
        return await _render_xlsx(artifact, blob)
    if ct.endswith("presentationml.presentation"):
        return await _render_pptx(artifact, blob)

    return {"type": "unsupported", "reason": f"no preview generator for {ct}"}


async def _render_docx(artifact: dict[str, Any], blob: bytes) -> dict[str, Any]:
    try:
        import mammoth  # type: ignore
    except ImportError:
        return {"type": "unsupported", "reason": "mammoth not installed; pip install mammoth"}
    out = preview_dir(artifact["id"], artifact["current_version_id"]) / "doc.html"
    if not out.exists():
        import io
        result = mammoth.convert_to_html(io.BytesIO(blob))
        out.write_text(result.value, encoding="utf-8")
    return {"type": "html", "content": out.read_text(encoding="utf-8")}


async def _render_xlsx(artifact: dict[str, Any], blob: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return {"type": "unsupported", "reason": "openpyxl not installed; pip install openpyxl"}
    import io
    wb = load_workbook(filename=io.BytesIO(blob), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell(v) for v in row])
            if len(rows) > 200:  # cap preview at 200 rows
                rows.append(["…"])
                break
        sheets.append({"name": ws.title, "rows": rows})
    return {"type": "sheet", "sheets": sheets}


def _cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


async def _render_pptx(artifact: dict[str, Any], blob: bytes) -> dict[str, Any]:
    """Render slides via libreoffice → PNG. Slow on first call, cached after."""
    if not shutil.which("soffice") and not shutil.which("libreoffice"):
        return {"type": "unsupported", "reason": "libreoffice (soffice) not installed"}
    out_dir = preview_dir(artifact["id"], artifact["current_version_id"])
    pdf_path = out_dir / "deck.pdf"
    if not pdf_path.exists():
        tmp_in = out_dir / "input.pptx"
        tmp_in.write_bytes(blob)
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        proc = await asyncio.create_subprocess_exec(
            soffice, "--headless", "--convert-to", "pdf",
            "--outdir", str(out_dir), str(tmp_in),
            stdout=asyncio.subprocess.DEVNULL, stderr=asyncio.subprocess.PIPE,
        )
        _, err = await proc.communicate()
        # libreoffice writes <stem>.pdf; rename to deck.pdf
        candidate = out_dir / (tmp_in.stem + ".pdf")
        if candidate.exists() and not pdf_path.exists():
            candidate.rename(pdf_path)
        if not pdf_path.exists():
            return {"type": "unsupported", "reason": f"libreoffice failed: {err.decode()[:200]}"}
    return {"type": "pdf", "url": f"/api/artifacts/{artifact['id']}/preview-pdf"}


async def get_preview_pdf(artifact_id: str, version_id: str) -> Path | None:
    """Return path to the cached PDF preview if it exists (PPTX flow)."""
    p = preview_dir(artifact_id, version_id) / "deck.pdf"
    return p if p.exists() else None
